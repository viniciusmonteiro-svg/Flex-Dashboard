"""
Build Reconciliation_ByChannel.xlsx matching the reference format exactly.
Tabs: Vendor x Channel | Channel Summary | Channel Mismatches
"""
import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────
EXCEL_SRC  = r'C:\Users\vmmon\Downloads\Untitled spreadsheet (8).xlsx'
DB_JSON    = r'scripts\full-recon-data.json'
OUT_PATH   = r'C:\Users\vmmon\Downloads\Reconciliation_ByChannel.xlsx'

QUARTERS = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'Q1 2026']
DB_INGESTED = {'Q3 2025', 'Q4 2025', 'Q1 2026'}

# ── Style helpers ───────────────────────────────────────────────────────────
def fill(h):  return PatternFill('solid', fgColor=h) if h else PatternFill()
def fnt(bold=False, color='000000', size=10, italic=False):
    return Font(name='Calibri', bold=bold, color=color, size=size, italic=italic)
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr(color='D0D0D0', style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def med_bdr():
    m = Side(style='medium', color='808080')
    t = Side(style='thin',   color='D0D0D0')
    return Border(left=t, right=t, top=m, bottom=m)

C_HDR    = '1F3864'   # dark navy
C_SUBHDR = '2E75B6'   # mid blue
C_REF_H  = '375623'   # dark green header  (Excel/Ref column group)
C_DB_H   = '1F3864'   # dark navy header   (DB column group)
C_DIF_H  = '7B2C2C'   # dark red header    (Diff column group)
C_REF_D  = 'E2EFDA'   # light green data
C_DB_D   = 'DEEAF1'   # light blue data
C_DIF_POS= 'C6EFCE'   # green diff
C_DIF_NEG= 'FFCCCC'   # red diff
C_DIF_ZRO= 'F2F2F2'   # grey diff near-zero
C_MATCH  = 'FFFFFF'
C_MISMATCH='FFF2CC'   # yellow mismatch row
C_UNCL   = 'FFCCCC'   # red uncl row
C_ALT    = 'F7FBFF'
C_TOTAL  = 'BDD7EE'
C_WHITE  = 'FFFFFF'
FMT_DOLLAR = '$#,##0.00'
FMT_DOLLAR0= '$#,##0'

def hcell(ws, r, c, text, bg=C_HDR, fg='FFFFFF', bold=True, size=10,
          h='center', wrap=False):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font      = fnt(bold=bold, color=fg, size=size)
    cell.fill      = fill(bg)
    cell.alignment = aln(h=h, wrap=wrap)
    cell.border    = bdr()
    return cell

def dcell(ws, r, c, val, bg=C_WHITE, bold=False, size=9, fmt=FMT_DOLLAR):
    cell = ws.cell(row=r, column=c)
    if val is None or val == 0:
        cell.value = None
    else:
        cell.value = val
        cell.number_format = fmt
    cell.font      = fnt(bold=bold, size=size)
    cell.fill      = fill(bg)
    cell.alignment = aln(h='right')
    cell.border    = bdr()
    return cell

def tcell(ws, r, c, text, bg=C_WHITE, bold=False, size=9, h='left',
          color='000000', italic=False):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font      = fnt(bold=bold, color=color, size=size, italic=italic)
    cell.fill      = fill(bg)
    cell.alignment = aln(h=h)
    cell.border    = bdr()
    return cell

def diff_fill(v):
    if   v is None or abs(v) < 0.5: return C_DIF_ZRO
    elif v > 0:  return C_DIF_POS
    else:        return C_DIF_NEG

# ── Load Excel source (Untitled 8) ─────────────────────────────────────────
src_wb = openpyxl.load_workbook(EXCEL_SRC, data_only=True)
src_ws = src_wb['Sheet1']
src_rows = list(src_ws.iter_rows(values_only=True))
src_hdr  = src_rows[0]
q_col    = {h: i for i, h in enumerate(src_hdr) if h in QUARTERS}
ch_col   = src_hdr.index('Marketing Channel')

# vendor → { excelChannel, quarters: { Q: $ } }
excel_vendors = {}
for r in src_rows[1:]:
    vendor  = r[0]
    channel = r[ch_col]
    if not vendor or not channel: continue
    if vendor not in excel_vendors:
        excel_vendors[vendor] = {'channel': channel, 'quarters': defaultdict(float)}
    for q, i in q_col.items():
        val = r[i]
        if val is not None:
            try: excel_vendors[vendor]['quarters'][q] += float(val)
            except: pass

# ── Load DB data ────────────────────────────────────────────────────────────
with open(DB_JSON, encoding='utf-8') as f:
    db = json.load(f)

db_rows   = db['rows']    # gl → vendor → { dbChannel, quarters: { Q: $ } }

# ── Build combined row list (GL × Vendor) ────────────────────────────────────
# Each item: (gl, vendor, excelChannel, dbChannel, status, {Q: (ref$, db$)})
combined = []
seen_excel = set()

for gl in sorted(db_rows.keys()):
    for vendor in sorted(db_rows[gl].keys(), key=str.lower):
        db_info    = db_rows[gl][vendor]
        db_channel = db_info['dbChannel']
        db_q       = db_info['quarters']   # Q → dollars (raw netsuite amount)

        ex_info     = excel_vendors.get(vendor)
        ex_channel  = ex_info['channel'] if ex_info else None
        ex_q        = ex_info['quarters'] if ex_info else {}

        seen_excel.add(vendor)

        # Status
        if ex_channel is None:
            status = 'Not in Excel'
        elif ex_channel == db_channel:
            status = 'MATCH'
        elif db_channel == 'Unclassified':
            status = 'Unclassified in DB'
        else:
            status = 'MISMATCH'

        # Per-quarter: ref$ = excel amount (or db if Q not in Excel)
        #              db$  = db amount
        q_data = {}
        for q in QUARTERS:
            ref = ex_q.get(q, 0) if ex_channel else 0
            dbv = db_q.get(q, 0)
            q_data[q] = (ref, dbv)

        combined.append({
            'gl': gl, 'vendor': vendor,
            'excelChannel': ex_channel or '—',
            'dbChannel': db_channel,
            'status': status,
            'quarters': q_data,
        })

# Vendors in Excel but NOT in DB at all
for vendor, ex_info in excel_vendors.items():
    found_in_db = any(vendor in db_rows[gl] for gl in db_rows)
    if not found_in_db:
        ex_q = ex_info['quarters']
        q_data = {}
        for q in QUARTERS:
            ref = ex_q.get(q, 0)
            q_data[q] = (ref, 0)
        combined.append({
            'gl': '—', 'vendor': vendor,
            'excelChannel': ex_info['channel'],
            'dbChannel': '—',
            'status': 'Excel Only',
            'quarters': q_data,
        })

# ── Channel Summary: Excel$ = sum of ref$ by excel channel
#                    DB$    = sum of db$  by db channel ─────────────────────
all_channels = sorted(set(
    [r['excelChannel'] for r in combined if r['excelChannel'] != '—'] +
    [r['dbChannel']    for r in combined if r['dbChannel']    != '—']
))

ch_excel_q = defaultdict(lambda: defaultdict(float))  # ch → Q → $
ch_db_q    = defaultdict(lambda: defaultdict(float))

for row in combined:
    for q in QUARTERS:
        ref, dbv = row['quarters'][q]
        if row['excelChannel'] != '—':
            ch_excel_q[row['excelChannel']][q] += ref
        if row['dbChannel'] != '—':
            ch_db_q[row['dbChannel']][q]        += dbv

# Mismatches only — rows where a channel was assigned in Excel but differs from DB
# (excludes "Not in Excel" which are pure DB rows with no reference channel)
mismatches = [r for r in combined
              if r['status'] in ('MISMATCH', 'Unclassified in DB', 'Excel Only')]

print(f"Total GL×Vendor rows: {len(combined)}")
print(f"  Mismatches/issues:  {len(mismatches)}")
print(f"  Channels:           {len(all_channels)}")

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def col_letter(n):   # 1-based
    return openpyxl.utils.get_column_letter(n)

# Column layout for Vendor x Channel and Mismatches:
# A=GL | B=Vendor | C=ExcelChannel | D=DBChannel | E=Status
# then per quarter: Ref$, DB$, Diff$  (3 cols each × 5 quarters = 15 cols)
# then Total Ref$, Total DB$, Total Diff$ (3 cols)
# Total: 5 + 15 + 3 = 23 columns  ← exactly matches reference
COL_GL    = 1
COL_VND   = 2
COL_ECH   = 3
COL_DCH   = 4
COL_STA   = 5
COL_Q_S   = 6   # start of quarter columns

def q_cols(qi):
    """Returns (ref_col, db_col, diff_col) for quarter index qi (0-based)"""
    base = COL_Q_S + qi * 3
    return base, base+1, base+2

COL_TREF  = COL_Q_S + len(QUARTERS)*3
COL_TDB   = COL_TREF + 1
COL_TDIFF = COL_TDB  + 1
LAST_COL  = COL_TDIFF

def build_header_row(ws, row):
    """Build the two-row header (group row + column names)"""
    # Row 1: group labels spanning quarter triplets + fixed cols
    headers_r1 = [
        (COL_GL,  1, 'GL Account'),
        (COL_VND, 1, 'Vendor / Entity'),
        (COL_ECH, 1, 'Excel Channel'),
        (COL_DCH, 1, 'DB Channel'),
        (COL_STA, 1, 'Channel Status'),
    ]
    for qi, q in enumerate(QUARTERS):
        rc, dc, xc = q_cols(qi)
        ws.merge_cells(start_row=row, start_column=rc, end_row=row, end_column=xc)
        hcell(ws, row, rc, q, bg=C_SUBHDR, size=9, wrap=False)
        for c in [dc, xc]:
            ws.cell(row=row, column=c).fill   = fill(C_SUBHDR)
            ws.cell(row=row, column=c).border = bdr()
    ws.merge_cells(start_row=row, start_column=COL_TREF,
                   end_row=row, end_column=COL_TDIFF)
    hcell(ws, row, COL_TREF, 'TOTAL', bg=C_HDR, size=9)
    for c in [COL_TDB, COL_TDIFF]:
        ws.cell(row=row, column=c).fill   = fill(C_HDR)
        ws.cell(row=row, column=c).border = bdr()

    for col, span, txt in headers_r1:
        if span > 1:
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col+span-1)
        hcell(ws, row, col, txt, size=9, wrap=True)

    # Row 2: column sub-labels
    r2 = row + 1
    for col, _, txt in headers_r1:
        ws.cell(row=r2, column=col).fill   = fill(C_HDR)
        ws.cell(row=r2, column=col).border = bdr()
        ws.cell(row=r2, column=col).value  = None

    for qi, q in enumerate(QUARTERS):
        rc, dc, xc = q_cols(qi)
        hcell(ws, r2, rc, f'{q} Ref $',  bg=C_REF_H, size=8)
        hcell(ws, r2, dc, f'{q} DB $',   bg=C_DB_H,  size=8)
        hcell(ws, r2, xc, f'{q} Diff $', bg=C_DIF_H, size=8)

    hcell(ws, r2, COL_TREF,  'Total Ref $',  bg=C_REF_H, size=8)
    hcell(ws, r2, COL_TDB,   'Total DB $',   bg=C_DB_H,  size=8)
    hcell(ws, r2, COL_TDIFF, 'Total Diff $', bg=C_DIF_H, size=8)

    ws.row_dimensions[row].height   = 18
    ws.row_dimensions[row+1].height = 22


def write_data_rows(ws, data_rows, start_row):
    """Write data rows, return next free row."""
    for ri, row in enumerate(data_rows):
        r = start_row + ri
        ws.row_dimensions[r].height = 15

        status = row['status']
        if   status == 'MATCH':             row_bg = C_MATCH
        elif status == 'MISMATCH':          row_bg = C_MISMATCH
        elif status == 'Unclassified in DB':row_bg = C_UNCL
        elif status == 'Excel Only':        row_bg = C_REF_D
        else:                               row_bg = C_ALT
        alt = C_ALT if ri % 2 else C_MATCH
        row_bg = row_bg  # status colour takes priority

        tcell(ws, r, COL_GL,  row['gl'],           bg=row_bg, size=8, color='444444')
        tcell(ws, r, COL_VND, row['vendor'],        bg=row_bg, size=9, bold=(status in ('MISMATCH','Unclassified in DB')))
        tcell(ws, r, COL_ECH, row['excelChannel'],  bg=row_bg, size=9)
        tcell(ws, r, COL_DCH, row['dbChannel'],     bg=row_bg, size=9)
        status_color = {'MATCH':'006100','MISMATCH':'9C5700',
                        'Unclassified in DB':'9C0006','Not in Excel':'555555',
                        'Excel Only':'375623'}.get(status,'000000')
        tcell(ws, r, COL_STA, status, bg=row_bg, size=9,
              color=status_color, bold=(status != 'Not in Excel'))

        t_ref = 0.0; t_db = 0.0
        for qi, q in enumerate(QUARTERS):
            rc, dc, xc = q_cols(qi)
            ref, dbv = row['quarters'][q]
            diff = dbv - ref
            t_ref += ref; t_db += dbv

            ref_bg = C_REF_D  if ref  else row_bg
            db_bg  = C_DB_D   if dbv  else row_bg
            diff_bg= diff_fill(diff)

            dcell(ws, r, rc, ref  or None, bg=ref_bg,  size=8)
            dcell(ws, r, dc, dbv  or None, bg=db_bg,   size=8)
            dcell(ws, r, xc, diff if abs(diff)>0.005 else None,
                  bg=diff_bg, size=8)

        t_diff = t_db - t_ref
        dcell(ws, r, COL_TREF,  t_ref  or None, bg=C_REF_D if t_ref else row_bg,  bold=True, size=8)
        dcell(ws, r, COL_TDB,   t_db   or None, bg=C_DB_D  if t_db  else row_bg,  bold=True, size=8)
        dcell(ws, r, COL_TDIFF, t_diff if abs(t_diff)>0.005 else None,
              bg=diff_fill(t_diff), bold=True, size=8)

    return start_row + len(data_rows)


def write_total_row(ws, r, label, data_rows):
    ws.row_dimensions[r].height = 18
    tcell(ws, r, COL_GL,  '',     bg=C_TOTAL, bold=True)
    tcell(ws, r, COL_VND, label,  bg=C_TOTAL, bold=True, size=10)
    tcell(ws, r, COL_ECH, '',     bg=C_TOTAL)
    tcell(ws, r, COL_DCH, '',     bg=C_TOTAL)
    tcell(ws, r, COL_STA, '',     bg=C_TOTAL)
    g_ref = 0.0; g_db = 0.0
    for qi, q in enumerate(QUARTERS):
        rc, dc, xc = q_cols(qi)
        q_ref = sum(row['quarters'][q][0] for row in data_rows)
        q_db  = sum(row['quarters'][q][1] for row in data_rows)
        q_diff= q_db - q_ref
        g_ref += q_ref; g_db += q_db
        dcell(ws, r, rc, q_ref  or None, bg=C_TOTAL, bold=True, size=9)
        dcell(ws, r, dc, q_db   or None, bg=C_TOTAL, bold=True, size=9)
        dcell(ws, r, xc, q_diff if abs(q_diff)>0.005 else None,
              bg=diff_fill(q_diff), bold=True, size=9)
    g_diff = g_db - g_ref
    dcell(ws, r, COL_TREF,  g_ref  or None, bg=C_TOTAL, bold=True)
    dcell(ws, r, COL_TDB,   g_db   or None, bg=C_TOTAL, bold=True)
    dcell(ws, r, COL_TDIFF, g_diff if abs(g_diff)>0.005 else None,
          bg=diff_fill(g_diff), bold=True)


def set_col_widths_detail(ws):
    ws.column_dimensions[col_letter(COL_GL)].width  = 38
    ws.column_dimensions[col_letter(COL_VND)].width = 34
    ws.column_dimensions[col_letter(COL_ECH)].width = 22
    ws.column_dimensions[col_letter(COL_DCH)].width = 28
    ws.column_dimensions[col_letter(COL_STA)].width = 22
    for qi in range(len(QUARTERS)):
        rc, dc, xc = q_cols(qi)
        for c in [rc, dc, xc]:
            ws.column_dimensions[col_letter(c)].width = 13
    for c in [COL_TREF, COL_TDB, COL_TDIFF]:
        ws.column_dimensions[col_letter(c)].width = 13


# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ── TAB 1: Vendor x Channel ──────────────────────────────────────────────────
ws1 = wb.create_sheet('Vendor x Channel')
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = 'F3'
set_col_widths_detail(ws1)

# Title
ws1.merge_cells(f'A1:{col_letter(LAST_COL)}1')
t = ws1['A1']
t.value     = 'Marketing Spend Reconciliation — All Channels, All Quarters'
t.font      = fnt(bold=True, color='FFFFFF', size=13)
t.fill      = fill(C_HDR)
t.alignment = aln()
ws1.row_dimensions[1].height = 24

build_header_row(ws1, 2)
next_row = write_data_rows(ws1, combined, start_row=4)
write_total_row(ws1, next_row, 'GRAND TOTAL', combined)

# ── TAB 2: Channel Summary ───────────────────────────────────────────────────
ws2 = wb.create_sheet('Channel Summary')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'B2'

# Column layout: A=Channel | per quarter: Excel$,DB$,Diff$ | Total Excel$,DB$,Diff$
# = 1 + 5*3 + 3 = 19 columns
S_COL_CH  = 1
S_COL_Q_S = 2
def s_q_cols(qi):
    base = S_COL_Q_S + qi * 3
    return base, base+1, base+2
S_COL_TREF  = S_COL_Q_S + len(QUARTERS)*3
S_COL_TDB   = S_COL_TREF + 1
S_COL_TDIFF = S_COL_TDB  + 1

ws2.column_dimensions['A'].width = 30
for qi in range(len(QUARTERS)):
    for c in s_q_cols(qi):
        ws2.column_dimensions[col_letter(c)].width = 14
for c in [S_COL_TREF, S_COL_TDB, S_COL_TDIFF]:
    ws2.column_dimensions[col_letter(c)].width = 14

# Title
last_s = col_letter(S_COL_TDIFF)
ws2.merge_cells(f'A1:{last_s}1')
t2 = ws2['A1']
t2.value = 'Channel Summary — Excel Reference vs DB Classification'
t2.font  = fnt(bold=True, color='FFFFFF', size=12)
t2.fill  = fill(C_HDR)
t2.alignment = aln()
ws2.row_dimensions[1].height = 22

# Group header
for qi, q in enumerate(QUARTERS):
    rc, dc, xc = s_q_cols(qi)
    ws2.merge_cells(start_row=2, start_column=rc, end_row=2, end_column=xc)
    hcell(ws2, 2, rc, q, bg=C_SUBHDR, size=9)
    for c in [dc, xc]:
        ws2.cell(row=2, column=c).fill   = fill(C_SUBHDR)
        ws2.cell(row=2, column=c).border = bdr()
ws2.merge_cells(start_row=2, start_column=S_COL_TREF, end_row=2, end_column=S_COL_TDIFF)
hcell(ws2, 2, S_COL_TREF, 'TOTAL', bg=C_HDR)
for c in [S_COL_TDB, S_COL_TDIFF]:
    ws2.cell(row=2, column=c).fill   = fill(C_HDR)
    ws2.cell(row=2, column=c).border = bdr()
hcell(ws2, 2, S_COL_CH, 'Channel', bg=C_HDR, h='left')
ws2.row_dimensions[2].height = 18

# Col sub-headers
hcell(ws2, 3, S_COL_CH, '', bg=C_HDR)
for qi, q in enumerate(QUARTERS):
    rc, dc, xc = s_q_cols(qi)
    hcell(ws2, 3, rc, f'{q} Excel $',  bg=C_REF_H, size=8)
    hcell(ws2, 3, dc, f'{q} DB $',     bg=C_DB_H,  size=8)
    hcell(ws2, 3, xc, f'{q} Diff $',   bg=C_DIF_H, size=8)
hcell(ws2, 3, S_COL_TREF,  'Total Excel $', bg=C_REF_H, size=8)
hcell(ws2, 3, S_COL_TDB,   'Total DB $',    bg=C_DB_H,  size=8)
hcell(ws2, 3, S_COL_TDIFF, 'Total Diff $',  bg=C_DIF_H, size=8)
ws2.row_dimensions[3].height = 20

# Data rows
DR = 4
for ci, ch in enumerate(all_channels):
    r   = DR + ci
    bg  = C_ALT if ci % 2 else C_WHITE
    ws2.row_dimensions[r].height = 17
    tcell(ws2, r, S_COL_CH, ch, bg=bg, bold=True, size=9)
    t_ex = 0.0; t_db2 = 0.0
    for qi, q in enumerate(QUARTERS):
        rc, dc, xc = s_q_cols(qi)
        ex_v  = ch_excel_q[ch].get(q, 0)
        db_v  = ch_db_q[ch].get(q, 0)
        diff  = db_v - ex_v
        t_ex += ex_v; t_db2 += db_v
        dcell(ws2, r, rc, ex_v or None, bg=C_REF_D if ex_v else bg, size=9)
        dcell(ws2, r, dc, db_v or None, bg=C_DB_D  if db_v else bg, size=9)
        dcell(ws2, r, xc, diff if abs(diff)>0.005 else None,
              bg=diff_fill(diff), size=9)
    t_diff = t_db2 - t_ex
    dcell(ws2, r, S_COL_TREF,  t_ex  or None, bg=C_REF_D if t_ex  else bg, bold=True)
    dcell(ws2, r, S_COL_TDB,   t_db2 or None, bg=C_DB_D  if t_db2 else bg, bold=True)
    dcell(ws2, r, S_COL_TDIFF, t_diff if abs(t_diff)>0.005 else None,
          bg=diff_fill(t_diff), bold=True)

# Grand total
GTROW = DR + len(all_channels)
ws2.row_dimensions[GTROW].height = 20
tcell(ws2, GTROW, S_COL_CH, 'GRAND TOTAL', bg=C_TOTAL, bold=True, size=10)
g_ex=0.0; g_db2=0.0
for qi, q in enumerate(QUARTERS):
    rc, dc, xc = s_q_cols(qi)
    q_ex  = sum(ch_excel_q[ch].get(q, 0) for ch in all_channels)
    q_db2 = sum(ch_db_q[ch].get(q,  0) for ch in all_channels)
    q_d   = q_db2 - q_ex
    g_ex += q_ex; g_db2 += q_db2
    dcell(ws2, GTROW, rc, q_ex  or None, bg=C_TOTAL, bold=True, size=9)
    dcell(ws2, GTROW, dc, q_db2 or None, bg=C_TOTAL, bold=True, size=9)
    dcell(ws2, GTROW, xc, q_d if abs(q_d)>0.005 else None,
          bg=diff_fill(q_d), bold=True, size=9)
g_d = g_db2 - g_ex
dcell(ws2, GTROW, S_COL_TREF,  g_ex  or None, bg=C_TOTAL, bold=True)
dcell(ws2, GTROW, S_COL_TDB,   g_db2 or None, bg=C_TOTAL, bold=True)
dcell(ws2, GTROW, S_COL_TDIFF, g_d if abs(g_d)>0.005 else None,
      bg=diff_fill(g_d), bold=True)

# ── TAB 3: Channel Mismatches ────────────────────────────────────────────────
ws3 = wb.create_sheet('Channel Mismatches')
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = 'F3'
set_col_widths_detail(ws3)

ws3.merge_cells(f'A1:{col_letter(LAST_COL)}1')
t3 = ws3['A1']
t3.value     = 'Channel Mismatches — Rows where Excel Channel ≠ DB Channel'
t3.font      = fnt(bold=True, color='FFFFFF', size=13)
t3.fill      = fill('7B2C2C')
t3.alignment = aln()
ws3.row_dimensions[1].height = 24

build_header_row(ws3, 2)
next_row3 = write_data_rows(ws3, mismatches, start_row=4)
write_total_row(ws3, next_row3, 'MISMATCH TOTAL', mismatches)

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
print(f'  Tab 1 "Vendor x Channel":   {len(combined)} rows')
print(f'  Tab 2 "Channel Summary":    {len(all_channels)} channels')
print(f'  Tab 3 "Channel Mismatches": {len(mismatches)} rows')
