"""
Build Trade Show Reconciliation Excel
  - Tab 1: Summary by Quarter
  - Tab 2-6: Per-quarter vendor detail (Excel vs DB vs Variance)
"""
import sys, json, math
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from collections import defaultdict

EXCEL_PATH   = r'C:\Users\vmmon\Downloads\Untitled spreadsheet (8).xlsx'
DB_JSON_PATH = r'scripts\recon-data.json'
OUT_PATH     = r'C:\Users\vmmon\Downloads\Trade Show Reconciliation.xlsx'

QUARTERS = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'Q1 2026']

# ── Colour palette ─────────────────────────────────────────────────────────
C_HEADER_BG  = '1F3864'   # dark navy
C_HEADER_FG  = 'FFFFFF'
C_SECTION_BG = 'D6E4F7'   # light blue
C_ALT_BG     = 'F2F7FD'   # very light blue (alternate rows)
C_POS_VAR    = 'C6EFCE'   # green fill  (DB > Excel)
C_NEG_VAR    = 'FFCCCC'   # red fill    (DB < Excel)
C_ZERO_VAR   = 'FFFFFF'
C_MISSING_BG = 'FFF2CC'   # yellow      (only in one source)
C_TITLE_BG   = '2E75B6'   # mid-blue for tab headers

def fill(hex_):  return PatternFill('solid', fgColor=hex_)
def font(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=False)
def left():   return Alignment(horizontal='left',   vertical='center')
def right():  return Alignment(horizontal='right',  vertical='center')
def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)
def bottom_border():
    s = Side(style='medium', color='808080')
    return Border(bottom=s)

FMT_DOLLAR  = '$#,##0.00'
FMT_DOLLAR0 = '$#,##0'

def set_dollar(cell, value, fmt=FMT_DOLLAR):
    cell.value = value if value is not None else None
    cell.number_format = fmt
    cell.alignment = right()
    cell.border = thin_border()

def header_cell(ws, row, col, text, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True,
                size=10, align=center):
    c = ws.cell(row=row, column=col, value=text)
    c.font  = font(bold=bold, color=fg, size=size)
    c.fill  = fill(bg)
    c.alignment = align()
    c.border = thin_border()
    return c

# ── Load Excel source data ─────────────────────────────────────────────────
src_wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
src_ws = src_wb['Sheet1']
src_rows = list(src_ws.iter_rows(values_only=True))
src_header = src_rows[0]  # ('Financial Row / Vendor','Marketing Channel','Notes / Reclass','Q1 2025',...)
q_col = {h: i for i, h in enumerate(src_header) if h in QUARTERS}

# Aggregate Excel by vendor per quarter  (some vendors appear on multiple rows)
excel_by_vendor_q = defaultdict(lambda: defaultdict(float))  # vendor → q → $
for r in src_rows[1:]:
    vendor = r[0]
    if not vendor:
        continue
    for q, i in q_col.items():
        val = r[i]
        if val is not None:
            try:
                excel_by_vendor_q[vendor][q] += float(val)
            except (TypeError, ValueError):
                pass

# Quarter totals from Excel
excel_q_totals = {}
for q in QUARTERS:
    excel_q_totals[q] = sum(v[q] for v in excel_by_vendor_q.values() if q in v)

# ── Load DB data ───────────────────────────────────────────────────────────
with open(DB_JSON_PATH, encoding='utf-8') as f:
    db = json.load(f)

db_by_vendor_q    = db['dbByVendorQ']      # vendor → { Q: dollars }
db_q_totals       = db['quarterTotals']    # Q → dollars
uncl_by_vendor_q  = db['unclByVendorQ']   # vendor → { channel, quarters: { Q: dollars } }

# Quarter totals from DB
def db_total(q):  return db_q_totals.get(q, 0)

# ── Build combined vendor list per quarter ─────────────────────────────────
all_vendors_q = {}  # q → sorted vendor list
for q in QUARTERS:
    vendors = set()
    for v, qmap in excel_by_vendor_q.items():
        if q in qmap:
            vendors.add(v)
    for v, qmap in db_by_vendor_q.items():
        if q in qmap:
            vendors.add(v)
    all_vendors_q[q] = sorted(vendors, key=lambda x: x.lower())

# ── Create workbook ────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ══════════════════════════════════════════════════════════════════════════
# TAB 1 — SUMMARY
# ══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('Summary')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 32

# Title
ws.merge_cells('A1:G1')
t = ws['A1']
t.value = 'Trade Show — Quarterly Reconciliation'
t.font  = font(bold=True, color=C_HEADER_FG, size=14)
t.fill  = fill(C_TITLE_BG)
t.alignment = center()
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:G2')
ws['A2'].value = 'Source: Attached Excel file vs Marketing Dashboard (DB)'
ws['A2'].font  = font(color='444444', size=9, italic=True)
ws['A2'].alignment = center()
ws.row_dimensions[2].height = 15

# Column headers
ROW = 4
for col, txt in enumerate(['Quarter', 'Excel (Source)', 'DB (Dashboard)',
                            'Variance (DB−Excel)', 'Variance %',
                            'DB excl. unclassified GL', 'Note'], 1):
    header_cell(ws, ROW, col, txt)
ws.row_dimensions[ROW].height = 20

# Data rows
DB_Q_MISSING_NOTE = 'Data not yet ingested in DB for this period'
for i, q in enumerate(QUARTERS):
    r = ROW + 1 + i
    ws.row_dimensions[r].height = 18
    bg = C_ALT_BG if i % 2 else C_ZERO_VAR

    excel_t = excel_q_totals.get(q, 0)
    db_t    = db_total(q)
    var     = db_t - excel_t
    pct     = (var / excel_t * 100) if excel_t else None

    # Uncl 70601/70602 for this quarter
    uncl_sum = sum(
        v['quarters'].get(q, 0)
        for v in uncl_by_vendor_q.values()
    )

    note = ''
    if abs(db_t) < 1 and excel_t > 0:
        note = DB_Q_MISSING_NOTE
    elif abs(var) < 0.01:
        note = 'Balanced ✓'

    # Col A: quarter label
    c = ws.cell(row=r, column=1, value=q)
    c.font = font(bold=True)
    c.fill = fill(bg)
    c.alignment = left()
    c.border = thin_border()

    # Col B: Excel total
    c = ws.cell(row=r, column=2)
    set_dollar(c, excel_t)
    c.fill = fill(bg)

    # Col C: DB total
    c = ws.cell(row=r, column=3)
    set_dollar(c, db_t)
    c.fill = fill(bg)
    if abs(db_t) < 1 and excel_t > 0:
        c.fill = fill(C_MISSING_BG)

    # Col D: Variance
    c = ws.cell(row=r, column=4)
    set_dollar(c, var)
    if abs(var) > 0.5:
        c.fill = fill(C_POS_VAR if var > 0 else C_NEG_VAR)
    else:
        c.fill = fill(C_ALT_BG if i % 2 else C_ZERO_VAR)

    # Col E: Variance %
    c = ws.cell(row=r, column=5)
    if pct is not None:
        c.value = pct / 100
        c.number_format = '0.00%'
    else:
        c.value = 'N/A'
    c.alignment = right()
    c.border = thin_border()
    c.fill = fill(bg)

    # Col F: DB incl unclassified GL (what would be TS if all 70601/70602 tagged)
    c = ws.cell(row=r, column=6)
    set_dollar(c, db_t + uncl_sum)
    c.fill = fill(bg)

    # Col G: Note
    c = ws.cell(row=r, column=7, value=note)
    c.font = font(color='444444', italic=bool(note))
    c.fill = fill(C_MISSING_BG if note == DB_Q_MISSING_NOTE else bg)
    c.alignment = left()
    c.border = thin_border()

# Grand total row
TOTAL_ROW = ROW + 1 + len(QUARTERS)
ws.row_dimensions[TOTAL_ROW].height = 20
ws.cell(row=TOTAL_ROW, column=1, value='TOTAL').font = font(bold=True, size=11)
ws.cell(row=TOTAL_ROW, column=1).fill  = fill(C_SECTION_BG)
ws.cell(row=TOTAL_ROW, column=1).alignment = left()
ws.cell(row=TOTAL_ROW, column=1).border = thin_border()

excel_grand = sum(excel_q_totals.values())
db_grand    = sum(db_total(q) for q in QUARTERS)
var_grand   = db_grand - excel_grand
uncl_grand  = sum(
    sum(v['quarters'].get(q, 0) for q in QUARTERS)
    for v in uncl_by_vendor_q.values()
)

for col, val in [(2, excel_grand), (3, db_grand), (4, var_grand),
                 (6, db_grand + uncl_grand)]:
    c = ws.cell(row=TOTAL_ROW, column=col)
    set_dollar(c, val)
    c.font = font(bold=True)
    c.fill = fill(C_SECTION_BG)

c = ws.cell(row=TOTAL_ROW, column=5)
pct_grand = (var_grand / excel_grand * 100) if excel_grand else None
c.value = (pct_grand / 100) if pct_grand else 'N/A'
c.number_format = '0.00%'
c.alignment = right()
c.border = thin_border()
c.font = font(bold=True)
c.fill = fill(C_SECTION_BG)

c = ws.cell(row=TOTAL_ROW, column=7, value='')
c.fill = fill(C_SECTION_BG)
c.border = thin_border()

# Legend
LR = TOTAL_ROW + 2
ws.cell(row=LR, column=1, value='Legend:').font = font(bold=True, size=9)
items = [
    (C_POS_VAR,   'DB > Excel (more in dashboard)'),
    (C_NEG_VAR,   'DB < Excel (less in dashboard)'),
    (C_MISSING_BG,'No DB data for this period'),
    ('F2F2F2',    '"DB excl. unclassified GL" = DB Trade Show + unclassified rows in GL 70601/70602'),
]
for j, (bg_hex, label) in enumerate(items):
    c = ws.cell(row=LR+j+1, column=1, value='')
    c.fill = fill(bg_hex)
    c.border = thin_border()
    ws.cell(row=LR+j+1, column=2, value=label).font = font(size=9, color='444444')

# ══════════════════════════════════════════════════════════════════════════
# TABS 2-6 — Per-Quarter Vendor Detail
# ══════════════════════════════════════════════════════════════════════════
for q in QUARTERS:
    ws = wb.create_sheet(q)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 44  # Vendor
    ws.column_dimensions['B'].width = 16  # Excel
    ws.column_dimensions['C'].width = 16  # DB
    ws.column_dimensions['D'].width = 16  # Variance
    ws.column_dimensions['E'].width = 14  # Variance %
    ws.column_dimensions['F'].width = 28  # Flag

    # Title
    ws.merge_cells('A1:F1')
    t = ws['A1']
    t.value = f'Trade Show Detail — {q}'
    t.font  = font(bold=True, color=C_HEADER_FG, size=13)
    t.fill  = fill(C_TITLE_BG)
    t.alignment = center()
    ws.row_dimensions[1].height = 24

    # Col headers
    for col, txt in enumerate(['Vendor', 'Excel ($)', 'DB — Dashboard ($)',
                                'Variance (DB−Excel)', 'Var %', 'Flag'], 1):
        header_cell(ws, 2, col, txt)
    ws.row_dimensions[2].height = 18

    vendors = all_vendors_q[q]
    data_start = 3
    db_q_missing = abs(db_total(q)) < 1 and excel_q_totals.get(q, 0) > 0

    for i, vendor in enumerate(vendors):
        r = data_start + i
        ws.row_dimensions[r].height = 16
        bg = C_ALT_BG if i % 2 else C_ZERO_VAR

        ex_amt = excel_by_vendor_q[vendor].get(q, 0) if vendor in excel_by_vendor_q else 0
        db_amt = db_by_vendor_q.get(vendor, {}).get(q, 0)
        var    = db_amt - ex_amt
        pct    = (var / ex_amt * 100) if ex_amt else None

        # Flag
        flag = ''
        if ex_amt == 0 and db_amt != 0:
            flag = '← In DB only'
        elif db_amt == 0 and ex_amt != 0:
            flag = '← In Excel only'
        if db_q_missing and ex_amt > 0:
            flag = '⚠ DB period not ingested'
        if abs(var) < 0.01 and ex_amt and db_amt:
            flag = '✓'

        # Vendor name
        c = ws.cell(row=r, column=1, value=vendor)
        c.font = font(size=9)
        c.fill = fill(bg)
        c.alignment = left()
        c.border = thin_border()

        # Excel amount
        c = ws.cell(row=r, column=2)
        set_dollar(c, ex_amt if ex_amt != 0 else None, FMT_DOLLAR)
        c.fill = fill(bg)
        c.font = font(size=9)
        if ex_amt == 0:
            c.value = '-'
            c.alignment = center()

        # DB amount
        c = ws.cell(row=r, column=3)
        set_dollar(c, db_amt if db_amt != 0 else None, FMT_DOLLAR)
        c.fill = fill(bg)
        c.font = font(size=9)
        if db_amt == 0:
            c.value = '-'
            c.alignment = center()
        if db_q_missing and ex_amt > 0 and db_amt == 0:
            c.fill = fill(C_MISSING_BG)

        # Variance
        c = ws.cell(row=r, column=4)
        if abs(var) < 0.005:
            c.value = '-'
            c.alignment = center()
            c.fill = fill(bg)
        else:
            set_dollar(c, var, FMT_DOLLAR)
            c.fill = fill(C_POS_VAR if var > 0 else C_NEG_VAR)
        c.font = font(size=9)
        c.border = thin_border()

        # Var %
        c = ws.cell(row=r, column=5)
        if pct is not None and abs(pct) > 0.01:
            c.value = pct / 100
            c.number_format = '0.0%'
        else:
            c.value = '-' if ex_amt else 'N/A'
        c.alignment = right() if pct else center()
        c.font = font(size=9)
        c.fill = fill(bg)
        c.border = thin_border()

        # Flag
        c = ws.cell(row=r, column=6, value=flag)
        c.font = font(size=9, color='444444', italic=flag not in ('✓', ''))
        c.fill = fill(C_MISSING_BG if '⚠' in flag else bg)
        c.alignment = left()
        c.border = thin_border()

    # ── Totals row ───────────────────────────────────────────────────────
    TROW = data_start + len(vendors)
    ws.row_dimensions[TROW].height = 18
    c = ws.cell(row=TROW, column=1, value='TOTAL')
    c.font  = font(bold=True)
    c.fill  = fill(C_SECTION_BG)
    c.alignment = left()
    c.border = thin_border()

    ex_tot = excel_q_totals.get(q, 0)
    db_tot = db_total(q)
    var_t  = db_tot - ex_tot

    for col, val in [(2, ex_tot), (3, db_tot), (4, var_t)]:
        c = ws.cell(row=TROW, column=col)
        set_dollar(c, val, FMT_DOLLAR)
        c.font = font(bold=True)
        c.fill = fill(C_SECTION_BG)
        if col == 4 and abs(var_t) > 0.5:
            c.fill = fill(C_POS_VAR if var_t > 0 else C_NEG_VAR)

    pct_t = (var_t / ex_tot * 100) if ex_tot else None
    c = ws.cell(row=TROW, column=5)
    c.value = (pct_t / 100) if pct_t else 'N/A'
    c.number_format = '0.0%'
    c.alignment = right()
    c.font = font(bold=True)
    c.fill = fill(C_SECTION_BG)
    c.border = thin_border()

    c = ws.cell(row=TROW, column=6, value='')
    c.fill  = fill(C_SECTION_BG)
    c.border = thin_border()

    # ── Unclassified 70601/70602 section for this quarter ────────────────
    uncl_vendors_here = {
        v: info for v, info in uncl_by_vendor_q.items()
        if q in info['quarters']
    }
    if uncl_vendors_here:
        ROW2 = TROW + 2
        ws.merge_cells(f'A{ROW2}:F{ROW2}')
        c = ws.cell(row=ROW2, column=1,
                    value=f'Unclassified rows in Trade Show GL accounts (70601/70602) — excluded from DB total above')
        c.font  = font(bold=True, color=C_HEADER_FG, size=9)
        c.fill  = fill('7F7F7F')
        c.alignment = left()

        header_cell(ws, ROW2+1, 1, 'Vendor', bg='595959', size=9)
        header_cell(ws, ROW2+1, 2, 'Amount ($)', bg='595959', size=9)
        header_cell(ws, ROW2+1, 3, 'Current Channel', bg='595959', size=9)
        for col in range(4, 7):
            ws.cell(row=ROW2+1, column=col).fill = fill('595959')
            ws.cell(row=ROW2+1, column=col).border = thin_border()

        for j, (vendor, info) in enumerate(sorted(uncl_vendors_here.items())):
            r2 = ROW2 + 2 + j
            amt = info['quarters'][q]
            ch  = info['channel']
            bg2 = C_ALT_BG if j % 2 else C_ZERO_VAR

            c = ws.cell(row=r2, column=1, value=vendor)
            c.font = font(size=9); c.fill = fill(bg2); c.alignment = left(); c.border = thin_border()

            c = ws.cell(row=r2, column=2)
            set_dollar(c, amt, FMT_DOLLAR)
            c.font = font(size=9); c.fill = fill(C_MISSING_BG)

            c = ws.cell(row=r2, column=3, value=ch)
            c.font = font(size=9, color='666666'); c.fill = fill(bg2); c.alignment = left(); c.border = thin_border()
            for col in range(4, 7):
                ws.cell(row=r2, column=col).fill = fill(bg2)
                ws.cell(row=r2, column=col).border = thin_border()

# ══════════════════════════════════════════════════════════════════════════
# Save
# ══════════════════════════════════════════════════════════════════════════
wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')

# Verify totals
print('\n=== Reconciliation Summary ===')
for q in QUARTERS:
    ex = excel_q_totals.get(q, 0)
    db = db_total(q)
    print(f'{q}:  Excel ${ex:>12,.2f}  |  DB ${db:>12,.2f}  |  Var ${db-ex:>10,.2f}')
print(f'{"TOTAL":6}:  Excel ${sum(excel_q_totals.values()):>12,.2f}  |  DB ${sum(db_total(q) for q in QUARTERS):>12,.2f}  |  Var ${sum(db_total(q) - excel_q_totals.get(q,0) for q in QUARTERS):>10,.2f}')
