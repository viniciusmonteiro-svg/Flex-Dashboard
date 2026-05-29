"""
Full Reconciliation Excel — All channels × all quarters.
  Tab 1 : Summary Matrix  (channel × quarter, DB vs Excel where available)
  Tabs 2+: Per-channel vendor detail
"""
import sys, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────────
EXCEL_SRC  = r'C:\Users\vmmon\Downloads\Untitled spreadsheet (8).xlsx'
DB_JSON    = r'scripts\db-all-channels.json'
EXCEL_JSON = r'scripts\excel-data.json'
OUT_PATH   = r'C:\Users\vmmon\Downloads\Marketing Spend Reconciliation.xlsx'

QUARTERS = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025', 'Q1 2026']

# Channel display order (most meaningful first)
CHANNEL_ORDER = [
    'Paid Search', 'Paid Social', 'Partner', 'SEO / Organic',
    'Review Sites', 'Referral', 'Email', 'Sales Development',
    'Trade Show', 'Other',
    'Unclassified', 'Do Not Tag (COGS/Non-S&M)',
]

# ── Colours ─────────────────────────────────────────────────────────────────
C_NAV  = '1F3864'   # dark navy — main headers
C_MID  = '2E75B6'   # mid-blue — tab titles
C_SECT = 'D6E4F7'   # light blue — section / totals
C_ALT  = 'F2F7FD'   # alternate row
C_WHT  = 'FFFFFF'
C_GRN  = 'C6EFCE'   # positive variance
C_RED  = 'FFCCCC'   # negative variance
C_YEL  = 'FFF2CC'   # warning / missing
C_GRY  = '595959'   # dark grey sub-header
C_LGRY = 'F2F2F2'   # light grey
C_DNT  = 'EDEDED'   # do-not-tag rows (neutral)

FMT_DOLLAR = '$#,##0.00'
FMT_PCT    = '0.0%'

def fill(h):  return PatternFill('solid', fgColor=h)
def fnt(bold=False, color='000000', size=10, italic=False):
    return Font(name='Arial', bold=bold, color=color, size=size, italic=italic)
def aln(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(color='BFBFBF', style='thin'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, row, col, text, bg=C_NAV, fg=C_WHT, bold=True, size=10, h='center'):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = fnt(bold=bold, color=fg, size=size)
    c.fill      = fill(bg)
    c.alignment = aln(h=h)
    c.border    = border()
    return c

def dcell(ws, row, col, value, bg=C_WHT, bold=False, fmt=FMT_DOLLAR, size=9):
    c = ws.cell(row=row, column=col)
    if value is None or value == 0:
        c.value         = None
        c.number_format = FMT_DOLLAR
    else:
        c.value         = value
        c.number_format = fmt
    c.font      = fnt(bold=bold, size=size)
    c.fill      = fill(bg)
    c.alignment = aln(h='right')
    c.border    = border()
    return c

def tcell(ws, row, col, text, bg=C_WHT, bold=False, size=9, h='left', color='000000', italic=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = fnt(bold=bold, color=color, size=size, italic=italic)
    c.fill      = fill(bg)
    c.alignment = aln(h=h)
    c.border    = border()
    return c

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# ── Load data ────────────────────────────────────────────────────────────────
with open(DB_JSON,    encoding='utf-8') as f: db    = json.load(f)
with open(EXCEL_JSON, encoding='utf-8') as f: exdata = json.load(f)

db_channels     = db['channels']        # list of channel names
db_channel_q    = db['channelQ']        # channel → Q → dollars
db_ch_vendor_q  = db['channelVendorQ']  # channel → vendor → Q → dollars

# Excel data (Trade Show only)
excel_ch_q      = exdata['channel_q_totals']   # channel → Q → dollars
excel_ch_vnd_q  = exdata['channel_vendor_q']   # channel → vendor → Q → dollars

# Build final ordered channel list: known order first, then any remaining
all_channels = [c for c in CHANNEL_ORDER if c in db_channels]
remaining    = [c for c in sorted(db_channels) if c not in CHANNEL_ORDER]
all_channels += remaining

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)

# ── Helper: quarter colour (for "not ingested" visual) ─────────────────────
DB_INGESTED = {'Q3 2025', 'Q4 2025', 'Q1 2026'}

def q_ingested(q): return q in DB_INGESTED

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — SUMMARY MATRIX
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet('Summary')
ws.sheet_view.showGridLines = False

# Column layout:
#   A: Channel  | B-F: DB per quarter | G: DB Total | H: Excel Total (TS only) | I: TS Variance
N_Q = len(QUARTERS)
col_db_start  = 2                    # B
col_db_end    = col_db_start + N_Q - 1   # F
col_db_total  = col_db_end + 1       # G
col_ex_total  = col_db_total + 1     # H
col_var       = col_ex_total + 1     # I

widths = {'A': 34}
for i, q in enumerate(QUARTERS):
    widths[chr(ord('A') + col_db_start - 1 + i)] = 14
widths[chr(ord('A') + col_db_total - 1)] = 14
widths[chr(ord('A') + col_ex_total - 1)] = 14
widths[chr(ord('A') + col_var - 1)]      = 16
set_col_widths(ws, widths)

# Title
last_col_letter = chr(ord('A') + col_var - 1)
ws.merge_cells(f'A1:{last_col_letter}1')
t = ws['A1']
t.value = 'Marketing Spend Reconciliation — All Channels by Quarter'
t.font  = fnt(bold=True, color=C_WHT, size=14)
t.fill  = fill(C_MID)
t.alignment = aln()
ws.row_dimensions[1].height = 28

ws.merge_cells(f'A2:{last_col_letter}2')
ws['A2'].value = 'DB = Marketing Dashboard (classified spend)  |  Excel = Attached reference file (Trade Show channel only)'
ws['A2'].font  = fnt(color='444444', size=9, italic=True)
ws['A2'].alignment = aln()
ws.row_dimensions[2].height = 14

# Sub-header groups row 3
ws.merge_cells(f'B3:{chr(ord("A")+col_db_end-1)}3')
c = ws['B3']
c.value = 'DB — Marketing Dashboard ($)'
c.font  = fnt(bold=True, color=C_WHT, size=9)
c.fill  = fill(C_NAV)
c.alignment = aln()
c.border = border()

hcell(ws, 3, col_db_total, 'DB Total', size=9)
hcell(ws, 3, col_ex_total, 'Excel Total\n(Trade Show ref.)', size=9, bg='4472C4')
hcell(ws, 3, col_var,      'Trade Show\nVariance (DB−Excel)', size=9, bg='4472C4')
ws.row_dimensions[3].height = 28

# Column headers row 4
hcell(ws, 4, 1, 'Channel', size=9)
for i, q in enumerate(QUARTERS):
    bg = C_NAV if q_ingested(q) else '7F7F7F'
    hcell(ws, 4, col_db_start + i, q, size=9, bg=bg)
hcell(ws, 4, col_db_total, 'All Qtrs', size=9)
hcell(ws, 4, col_ex_total, 'All Qtrs', size=9, bg='4472C4')
hcell(ws, 4, col_var,      'All Qtrs', size=9, bg='4472C4')
ws.row_dimensions[4].height = 18

# Data rows
ROW = 5
for ci, channel in enumerate(all_channels):
    ws.row_dimensions[ROW+ci].height = 17

    # Muted background for Do Not Tag
    is_dnt    = channel == 'Do Not Tag (COGS/Non-S&M)'
    is_uncl   = channel == 'Unclassified'
    row_bg    = C_DNT if is_dnt else (C_ALT if ci % 2 else C_WHT)
    ch_color  = '888888' if is_dnt else '000000'

    c = ws.cell(row=ROW+ci, column=1, value=channel)
    c.font      = fnt(bold=not is_dnt, color=ch_color, size=9, italic=is_dnt)
    c.fill      = fill(row_bg)
    c.alignment = aln(h='left')
    c.border    = border()

    db_total_ch = 0
    for i, q in enumerate(QUARTERS):
        val = db_channel_q.get(channel, {}).get(q, 0)
        db_total_ch += val
        col = col_db_start + i
        bg  = row_bg if q_ingested(q) else C_YEL
        c   = dcell(ws, ROW+ci, col, val or None, bg=bg, size=9)
        c.font = fnt(color=ch_color, size=9, italic=is_dnt)

    # DB total
    c = dcell(ws, ROW+ci, col_db_total, db_total_ch or None, bg=row_bg, size=9)
    c.font = fnt(bold=True, color=ch_color, size=9, italic=is_dnt)

    # Excel total (Trade Show only)
    ex_total_ch = sum(excel_ch_q.get(channel, {}).get(q, 0) for q in QUARTERS)
    c = dcell(ws, ROW+ci, col_ex_total, ex_total_ch or None,
              bg='EBF3FB' if ex_total_ch else row_bg, size=9)

    # Variance (only meaningful for Trade Show)
    var = db_total_ch - ex_total_ch
    if ex_total_ch or (channel == 'Trade Show'):
        vbg = C_GRN if var > 0.5 else (C_RED if var < -0.5 else row_bg)
        c = dcell(ws, ROW+ci, col_var, var or None, bg=vbg, size=9)
    else:
        c = ws.cell(row=ROW+ci, column=col_var, value='—')
        c.fill = fill(row_bg); c.border = border()
        c.alignment = aln(); c.font = fnt(color='AAAAAA', size=9)

# Totals row
TROW = ROW + len(all_channels)
ws.row_dimensions[TROW].height = 20
tcell(ws, TROW, 1, 'TOTAL', bg=C_SECT, bold=True, size=10)

grand_db = 0; grand_ex = 0
for i, q in enumerate(QUARTERS):
    val = sum(db_channel_q.get(ch, {}).get(q, 0) for ch in all_channels)
    grand_db += val
    dcell(ws, TROW, col_db_start+i, val or None, bg=C_SECT, bold=True, size=9)

dcell(ws, TROW, col_db_total, grand_db or None, bg=C_SECT, bold=True)

grand_ex = sum(sum(excel_ch_q.get(ch, {}).get(q, 0) for q in QUARTERS) for ch in all_channels)
dcell(ws, TROW, col_ex_total, grand_ex or None, bg=C_SECT, bold=True)

grand_var = grand_db - grand_ex  # only TS has excel data so this = TS variance
dcell(ws, TROW, col_var,
      grand_var if abs(grand_var) > 0.5 else None,
      bg=C_GRN if grand_var > 0.5 else (C_RED if grand_var < -0.5 else C_SECT),
      bold=True)

# Legend
LR = TROW + 2
tcell(ws, LR, 1, 'Notes:', bold=True, size=9)
notes = [
    (C_YEL,  'Quarter not yet ingested into the DB — DB figure will be $0'),
    (C_GRN,  'Variance: DB > Excel (more in dashboard than reference)'),
    (C_RED,  'Variance: DB < Excel (less in dashboard than reference)'),
    (C_DNT,  '"Do Not Tag (COGS/Non-S&M)" = COGS/overhead GL accounts, excluded from S&M reporting'),
    (C_LGRY, 'Excel column only populated for Trade Show (the provided reference file)'),
]
for j, (bg, txt) in enumerate(notes):
    r = LR + 1 + j
    c = ws.cell(row=r, column=1, value='')
    c.fill = fill(bg); c.border = border()
    ws.cell(row=r, column=2, value=txt).font = fnt(size=9, color='444444')
    ws.merge_cells(f'B{r}:{last_col_letter}{r}')

# ══════════════════════════════════════════════════════════════════════════════
# TABS per channel — vendor detail
# ══════════════════════════════════════════════════════════════════════════════
for channel in all_channels:
    # Shorten tab name for Excel's 31-char limit; strip invalid chars (/ \ ? * [ ])
    import re
    tab_name = re.sub(r'[/\\?*\[\]]', '-', channel)[:31]
    ws = wb.create_sheet(tab_name)
    ws.sheet_view.showGridLines = False

    # Column layout: A=Vendor | B-F=DB qtrs | G=DB Total | H-L=Excel qtrs | M=Excel Total | N=Variance | O=Var%
    COL_DB_S  = 2
    COL_DB_E  = COL_DB_S + N_Q - 1
    COL_DB_T  = COL_DB_E + 1
    COL_EX_S  = COL_DB_T + 1
    COL_EX_E  = COL_EX_S + N_Q - 1
    COL_EX_T  = COL_EX_E + 1
    COL_VAR   = COL_EX_T + 1
    COL_VARP  = COL_VAR + 1
    COL_FLAG  = COL_VARP + 1

    set_col_widths(ws, {
        'A': 42,
        **{chr(ord('A')+COL_DB_S-1+i): 13 for i in range(N_Q)},
        chr(ord('A')+COL_DB_T-1): 13,
        **{chr(ord('A')+COL_EX_S-1+i): 13 for i in range(N_Q)},
        chr(ord('A')+COL_EX_T-1): 13,
        chr(ord('A')+COL_VAR-1): 14,
        chr(ord('A')+COL_VARP-1): 9,
        chr(ord('A')+COL_FLAG-1): 26,
    })

    last_c = chr(ord('A') + COL_FLAG - 1)

    # Title
    ws.merge_cells(f'A1:{last_c}1')
    t = ws['A1']
    t.value = f'{channel} — Quarterly Vendor Detail'
    t.font  = fnt(bold=True, color=C_WHT, size=13)
    t.fill  = fill(C_MID)
    t.alignment = aln()
    ws.row_dimensions[1].height = 24

    # Group header row 2
    ws.merge_cells(f'{chr(ord("A")+COL_DB_S-1)}2:{chr(ord("A")+COL_DB_T-1)}2')
    c = ws[f'{chr(ord("A")+COL_DB_S-1)}2']
    c.value = 'DB — Marketing Dashboard ($)'; c.font = fnt(bold=True, color=C_WHT, size=9)
    c.fill = fill(C_NAV); c.alignment = aln(); c.border = border()

    has_excel = channel in excel_ch_q
    ex_label  = 'Excel Reference ($)' if has_excel else 'Excel Reference (no data for this channel)'
    ex_bg     = '4472C4' if has_excel else '7F7F7F'
    ws.merge_cells(f'{chr(ord("A")+COL_EX_S-1)}2:{chr(ord("A")+COL_EX_T-1)}2')
    c = ws[f'{chr(ord("A")+COL_EX_S-1)}2']
    c.value = ex_label; c.font = fnt(bold=True, color=C_WHT, size=9)
    c.fill = fill(ex_bg); c.alignment = aln(); c.border = border()

    for col in [COL_VAR, COL_VARP, COL_FLAG]:
        ws.cell(row=2, column=col).fill  = fill(C_NAV)
        ws.cell(row=2, column=col).border = border()

    ws.row_dimensions[2].height = 18

    # Column headers row 3
    hcell(ws, 3, 1, 'Vendor', size=9, h='left')
    for i, q in enumerate(QUARTERS):
        bg = C_NAV if q_ingested(q) else '7F7F7F'
        hcell(ws, 3, COL_DB_S+i, q, size=9, bg=bg)
    hcell(ws, 3, COL_DB_T, 'Total', size=9)
    for i, q in enumerate(QUARTERS):
        hcell(ws, 3, COL_EX_S+i, q, size=9, bg=ex_bg)
    hcell(ws, 3, COL_EX_T, 'Total', size=9, bg=ex_bg)
    hcell(ws, 3, COL_VAR, 'Variance\n(DB−Excel)', size=9)
    hcell(ws, 3, COL_VARP, 'Var%', size=9)
    hcell(ws, 3, COL_FLAG, 'Flag', size=9)
    ws.row_dimensions[3].height = 24

    # Combine vendors from both sources
    db_vnd_q  = db_ch_vendor_q.get(channel, {})
    ex_vnd_q  = excel_ch_vnd_q.get(channel, {})
    all_vnd   = sorted(set(list(db_vnd_q.keys()) + list(ex_vnd_q.keys())),
                        key=lambda x: x.lower())

    is_dnt = channel == 'Do Not Tag (COGS/Non-S&M)'

    DATA_START = 4
    for ri, vendor in enumerate(all_vnd):
        r   = DATA_START + ri
        bg  = C_DNT if is_dnt else (C_ALT if ri % 2 else C_WHT)
        ws.row_dimensions[r].height = 15

        tcell(ws, r, 1, vendor, bg=bg, size=9,
              color='888888' if is_dnt else '000000',
              italic=is_dnt)

        db_tot_v = 0
        for i, q in enumerate(QUARTERS):
            val   = db_vnd_q.get(vendor, {}).get(q, 0)
            db_tot_v += val
            qbg = bg if q_ingested(q) else C_YEL
            dcell(ws, r, COL_DB_S+i, val or None, bg=qbg, size=9)
        dcell(ws, r, COL_DB_T, db_tot_v or None, bg=bg, bold=True, size=9)

        ex_tot_v = 0
        for i, q in enumerate(QUARTERS):
            val = ex_vnd_q.get(vendor, {}).get(q, 0)
            ex_tot_v += val
            dcell(ws, r, COL_EX_S+i, val or None, bg=bg if has_excel else C_LGRY, size=9)
        dcell(ws, r, COL_EX_T, ex_tot_v or None, bg=bg if has_excel else C_LGRY,
              bold=True, size=9)

        var  = db_tot_v - ex_tot_v
        pct  = (var / ex_tot_v) if ex_tot_v else None
        vbg  = (C_GRN if var > 0.5 else (C_RED if var < -0.5 else bg)) if has_excel else bg

        if has_excel and abs(var) > 0.005:
            dcell(ws, r, COL_VAR, var, bg=vbg, size=9)
        else:
            c = ws.cell(row=r, column=COL_VAR, value='—' if not has_excel else None)
            c.fill = fill(bg); c.border = border()
            c.alignment = aln(); c.font = fnt(color='AAAAAA', size=9)

        c = ws.cell(row=r, column=COL_VARP)
        if pct is not None and has_excel and abs(pct) > 0.0001:
            c.value = pct; c.number_format = FMT_PCT
        else:
            c.value = '—'
        c.alignment = aln(h='right'); c.border = border()
        c.fill = fill(bg); c.font = fnt(size=9, color='AAAAAA' if c.value == '—' else '000000')

        flag = ''
        if db_tot_v == 0 and ex_tot_v != 0:
            flag = '← Excel only'
        elif ex_tot_v == 0 and db_tot_v != 0 and has_excel:
            flag = '← DB only'
        elif has_excel and abs(var) < 0.01 and ex_tot_v:
            flag = '✓ Balanced'
        tcell(ws, r, COL_FLAG, flag, bg=bg, size=9,
              color='888888', italic=flag not in ('✓ Balanced', ''))

    # Totals row
    TROW = DATA_START + len(all_vnd)
    ws.row_dimensions[TROW].height = 18
    tcell(ws, TROW, 1, 'TOTAL', bg=C_SECT, bold=True, size=10)

    db_grand_v = 0; ex_grand_v = 0
    for i, q in enumerate(QUARTERS):
        val = sum(db_vnd_q.get(v, {}).get(q, 0) for v in all_vnd)
        db_grand_v += val
        dcell(ws, TROW, COL_DB_S+i, val or None, bg=C_SECT, bold=True, size=9)
    dcell(ws, TROW, COL_DB_T, db_grand_v or None, bg=C_SECT, bold=True)

    for i, q in enumerate(QUARTERS):
        val = sum(ex_vnd_q.get(v, {}).get(q, 0) for v in all_vnd)
        ex_grand_v += val
        dcell(ws, TROW, COL_EX_S+i, val or None, bg=C_SECT, bold=True, size=9)
    dcell(ws, TROW, COL_EX_T, ex_grand_v or None, bg=C_SECT, bold=True)

    var_grand = db_grand_v - ex_grand_v
    dcell(ws, TROW, COL_VAR, var_grand if abs(var_grand) > 0.5 else None,
          bg=C_GRN if var_grand > 0.5 else (C_RED if var_grand < -0.5 else C_SECT), bold=True)

    pct_grand = (var_grand / ex_grand_v) if ex_grand_v else None
    c = ws.cell(row=TROW, column=COL_VARP)
    if pct_grand is not None and has_excel:
        c.value = pct_grand; c.number_format = FMT_PCT
    else:
        c.value = '—'
    c.alignment = aln(h='right'); c.border = border()
    c.fill = fill(C_SECT); c.font = fnt(bold=True, size=9)

    c = ws.cell(row=TROW, column=COL_FLAG, value='')
    c.fill = fill(C_SECT); c.border = border()

# ══════════════════════════════════════════════════════════════════════════════
wb.save(OUT_PATH)
print(f'Saved: {OUT_PATH}')
print(f'Tabs: Summary + {len(all_channels)} channel tabs = {1+len(all_channels)} total')
